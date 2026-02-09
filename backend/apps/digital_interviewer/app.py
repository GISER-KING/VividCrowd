"""
数字面试官应用 - FastAPI路由
"""
import os
import hashlib
import uuid
from typing import List, Optional
from fastapi import APIRouter, UploadFile, File, Form, HTTPException, WebSocket, WebSocketDisconnect, Depends
from fastapi.responses import JSONResponse
from sqlalchemy.orm import Session
from datetime import datetime

from backend.core.database import get_digital_interviewer_db
from backend.models.db_models import (
    InterviewerProfile,
    InterviewerProfileRegistry,
    InterviewSession,
    InterviewRound,
    InterviewEvaluation,
    InterviewKnowledge,
    InterviewExperienceSet
)
from backend.apps.digital_interviewer.services.profile_parser import InterviewerProfileParser
from backend.apps.digital_interviewer.services.experience_service import ExperienceService

# 创建路由器
router = APIRouter(prefix="/digital-interviewer", tags=["数字面试官"])

# 文件上传目录
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data", "interviewer_uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

VIDEO_DIR = os.path.join(UPLOAD_DIR, "videos")
os.makedirs(VIDEO_DIR, exist_ok=True)

PROFILE_DIR = os.path.join(UPLOAD_DIR, "profiles")
os.makedirs(PROFILE_DIR, exist_ok=True)

# 面经文件上传目录
EXPERIENCE_DIR = os.path.join(UPLOAD_DIR, "experiences")
os.makedirs(EXPERIENCE_DIR, exist_ok=True)

# 简历文件上传目录
RESUME_DIR = os.path.join(UPLOAD_DIR, "resumes")
os.makedirs(RESUME_DIR, exist_ok=True)

# 虚拟人形象目录（独立的虚拟人形象库）
DIGITAL_HUMANS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data", "digital_humans")
os.makedirs(DIGITAL_HUMANS_DIR, exist_ok=True)


def scan_all_digital_humans():
    """
    启动时扫描所有虚拟人形象
    """
    from loguru import logger
    from backend.core.database import get_digital_interviewer_db
    from backend.models.db_models import DigitalHuman

    logger.info("🎭 开始扫描虚拟人形象库...")

    db = next(get_digital_interviewer_db())
    try:
        if not os.path.exists(DIGITAL_HUMANS_DIR):
            logger.warning(f"⚠️ 虚拟人形象目录不存在: {DIGITAL_HUMANS_DIR}")
            return 0

        total_count = 0

        # 遍历所有子目录
        for human_name in os.listdir(DIGITAL_HUMANS_DIR):
            human_path = os.path.join(DIGITAL_HUMANS_DIR, human_name)

            # 只处理目录
            if not os.path.isdir(human_path):
                continue

            logger.info(f"🎭 扫描虚拟人形象: {human_name}")

            # 检查4个必需的视频文件
            video_files = {
                'video_idle': 'idle.mp4',
                'video_speaking': 'speaking.mp4',
                'video_listening': 'listening.mp4',
                'video_thinking': 'thinking.mp4'
            }

            videos = {}
            all_found = True
            for field_name, filename in video_files.items():
                file_path = os.path.join(human_path, filename)
                if os.path.exists(file_path):
                    # 返回相对路径，便于前端访问
                    relative_path = f"/data/digital_humans/{human_name}/{filename}"
                    videos[field_name] = relative_path
                    logger.debug(f"  ✅ {filename}")
                else:
                    logger.warning(f"  ⚠️ 缺少 {filename}")
                    all_found = False

            if not all_found:
                logger.warning(f"  ⚠️ 虚拟人形象 {human_name} 视频不完整，跳过")
                continue

            # 检查数据库中是否已存在该形象
            existing = db.query(DigitalHuman).filter_by(name=human_name).first()

            if existing:
                # 更新现有记录
                for key, value in videos.items():
                    setattr(existing, key, value)
                existing.is_active = True
                logger.info(f"  🔄 更新虚拟人形象: {human_name}")
            else:
                # 创建新记录
                # 自动生成显示名称（将下划线替换为空格并首字母大写）
                display_name = human_name.replace('_', ' ').title()

                # 尝试从名称中提取性别和风格
                gender = None
                style = None
                if 'male' in human_name.lower():
                    gender = 'male'
                elif 'female' in human_name.lower():
                    gender = 'female'

                if 'formal' in human_name.lower():
                    style = 'formal'
                elif 'casual' in human_name.lower():
                    style = 'casual'
                elif 'tech' in human_name.lower():
                    style = 'tech'

                human = DigitalHuman(
                    name=human_name,
                    display_name=display_name,
                    gender=gender,
                    style=style,
                    is_active=True,
                    is_default=(human_name == 'default'),
                    **videos
                )
                db.add(human)
                logger.info(f"  ✨ 新增虚拟人形象: {display_name}")

            total_count += 1

        db.commit()
        logger.info(f"✅ 扫描完成，共找到 {total_count} 个虚拟人形象")
        return total_count

    except Exception as e:
        logger.error(f"❌ 扫描失败: {str(e)}")
        logger.exception(e)
        return 0
    finally:
        db.close()


@router.get("/interviewers")
async def get_interviewers(db: Session = Depends(get_digital_interviewer_db)):
    """获取所有面试官列表"""
    interviewers = db.query(InterviewerProfile).all()
    return {"interviewers": [interviewer.to_dict() for interviewer in interviewers]}


@router.get("/digital-humans")
async def get_digital_humans(db: Session = Depends(get_digital_interviewer_db)):
    """获取所有可用的虚拟人形象"""
    from backend.models.db_models import DigitalHuman

    humans = db.query(DigitalHuman).filter_by(is_active=True).all()

    return {
        "digital_humans": [human.to_dict() for human in humans]
    }


@router.post("/interviewers/upload")
async def upload_interviewer_profile(
    file: UploadFile = File(...),
    db: Session = Depends(get_digital_interviewer_db)
):
    """上传面试官画像文件"""
    from loguru import logger

    try:
        logger.info(f"📤 开始上传面试官画像: {file.filename}")

        # 计算文件哈希
        content = await file.read()
        file_hash = hashlib.md5(content).hexdigest()
        logger.info(f"📝 文件哈希: {file_hash}")

        # 检查是否已导入
        existing = db.query(InterviewerProfileRegistry).filter_by(file_hash=file_hash).first()
        if existing:
            logger.info(f"⚠️ 文件已存在，跳过导入")
            return {"message": "该文件已导入", "interviewer_id": existing.interviewer_profile_id}

        # 保存文件
        file_path = os.path.join(PROFILE_DIR, file.filename)
        with open(file_path, 'wb') as f:
            f.write(content)
        logger.info(f"💾 文件已保存到: {file_path}")

        # 解析画像
        logger.info(f"🔍 开始解析画像文件...")
        parser = InterviewerProfileParser()
        profile_data = await parser.parse_file(file_path)
        logger.info(f"✅ 解析完成，提取到的数据: {profile_data.keys()}")
        logger.info(f"📋 姓名: {profile_data.get('name', 'N/A')}")
        logger.info(f"📋 专业领域: {profile_data.get('expertise_areas', 'N/A')}")

        # 生成系统提示词
        system_prompt = await parser.generate_system_prompt(profile_data)
        profile_data['system_prompt'] = system_prompt
        logger.info(f"💬 系统提示词已生成 (长度: {len(system_prompt)})")

        # 保存到数据库
        interviewer = InterviewerProfile(**profile_data)
        db.add(interviewer)
        db.commit()
        db.refresh(interviewer)
        logger.info(f"✅ 面试官已保存到数据库，ID: {interviewer.id}")

        # 注册文件
        registry = InterviewerProfileRegistry(
            filename=file.filename,
            file_hash=file_hash,
            interviewer_profile_id=interviewer.id,
            interviewer_name=interviewer.name
        )
        db.add(registry)
        db.commit()
        logger.info(f"✅ 文件已注册")

        return {"message": "上传成功", "interviewer": interviewer.to_dict()}

    except Exception as e:
        logger.error(f"❌ 上传失败: {str(e)}")
        logger.exception(e)
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/interviewers/{interviewer_id}")
async def delete_interviewer(
    interviewer_id: int,
    db: Session = Depends(get_digital_interviewer_db)
):
    """删除面试官"""
    try:
        interviewer = db.query(InterviewerProfile).filter_by(id=interviewer_id).first()
        if not interviewer:
            raise HTTPException(status_code=404, detail="面试官不存在")

        db.delete(interviewer)
        db.commit()

        return {"message": "删除成功"}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/sessions/start")
async def start_interview_session(
    interviewer_id: int = Form(...),
    interview_type: str = Form(...),
    candidate_name: str = Form(None),
    digital_human_id: int = Form(None),
    experience_set_ids: str = Form(None),  # JSON数组字符串，如 "[1,2,3]"
    experience_mode: str = Form("none"),   # none/reference/strict/mixed
    max_rounds: int = Form(5),             # 最大面试轮数，默认5轮
    db: Session = Depends(get_digital_interviewer_db)
):
    """开始面试会话"""
    from backend.models.db_models import DigitalHuman
    import json as json_module

    try:
        interviewer = db.query(InterviewerProfile).filter_by(id=interviewer_id).first()
        if not interviewer:
            raise HTTPException(status_code=404, detail="面试官不存在")

        # 验证虚拟人形象
        digital_human = None
        if digital_human_id:
            digital_human = db.query(DigitalHuman).filter_by(id=digital_human_id).first()
            if not digital_human:
                raise HTTPException(status_code=404, detail="虚拟人形象不存在")

        # 解析面经集ID列表
        parsed_set_ids = []
        if experience_set_ids and experience_set_ids.strip():
            try:
                parsed_set_ids = json_module.loads(experience_set_ids)
                if not isinstance(parsed_set_ids, list):
                    parsed_set_ids = []
            except:
                parsed_set_ids = []

        # 创建会话
        session_id = str(uuid.uuid4())
        session = InterviewSession(
            session_id=session_id,
            interviewer_profile_id=interviewer_id,
            digital_human_id=digital_human_id,
            interviewer_name=interviewer.name,
            interviewer_title=interviewer.title,
            interview_type=interview_type,
            candidate_name=candidate_name,
            max_rounds=max_rounds,
            status="in_progress"
        )

        db.add(session)
        db.commit()
        db.refresh(session)

        # 返回会话信息和虚拟人形象视频路径
        response_data = {
            "message": "会话创建成功",
            "session_id": session.session_id,
            "session": session.to_dict(),
            "experience_set_ids": parsed_set_ids,
            "experience_mode": experience_mode,
            "max_rounds": max_rounds
        }

        if digital_human:
            # 构建完整的视频URL（包含后端服务器地址）
            base_url = "http://localhost:8001"  # 后端服务器地址（注意是8001端口）
            response_data["digital_human_videos"] = {
                "idle": f"{base_url}{digital_human.video_idle}",
                "speaking": f"{base_url}{digital_human.video_speaking}",
                "listening": f"{base_url}{digital_human.video_listening}",
                "thinking": f"{base_url}{digital_human.video_thinking}"
            }

        return response_data

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sessions/{session_id}")
async def get_interview_session(
    session_id: str,
    db: Session = Depends(get_digital_interviewer_db)
):
    """获取面试会话详情"""
    session = db.query(InterviewSession).filter_by(session_id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    return {"session": session.to_dict()}


@router.get("/sessions/{session_id}/rounds")
async def get_interview_rounds(
    session_id: str,
    db: Session = Depends(get_digital_interviewer_db)
):
    """获取面试轮次数据"""
    from backend.models.db_models import InterviewRound

    session = db.query(InterviewSession).filter_by(session_id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    rounds = db.query(InterviewRound).filter_by(session_id=session.id).order_by(InterviewRound.round_number).all()
    return {"rounds": [r.to_dict() for r in rounds]}


@router.get("/sessions/{session_id}/evaluation")
async def get_interview_evaluation(
    session_id: str,
    db: Session = Depends(get_digital_interviewer_db)
):
    """获取面试评估结果"""
    from backend.models.db_models import InterviewEvaluation, InterviewRound

    session = db.query(InterviewSession).filter_by(session_id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    # 尝试获取已保存的评估
    evaluation = db.query(InterviewEvaluation).filter_by(session_id=session.id).first()

    if evaluation:
        return {"evaluation": evaluation.to_dict()}

    # 如果没有保存的评估，从轮次数据中汇总
    rounds = db.query(InterviewRound).filter_by(session_id=session.id).all()

    if not rounds:
        # 没有任何轮次数据，返回空评估
        return {
            "evaluation": {
                "technical_score": 0,
                "communication_score": 0,
                "problem_solving_score": 0,
                "cultural_fit_score": 0,
                "total_score": 0,
                "performance_level": "未完成",
                "strengths": [],
                "weaknesses": [],
                "suggestions": ["面试未完成，无法生成评估"]
            }
        }

    # 从轮次评估数据中汇总
    total_technical = 0
    total_communication = 0
    total_problem_solving = 0
    total_cultural_fit = 0
    count = 0

    for r in rounds:
        if r.evaluation_data:
            total_technical += r.evaluation_data.get("technical_score", 0)
            total_communication += r.evaluation_data.get("communication_score", 0)
            total_problem_solving += r.evaluation_data.get("problem_solving_score", 0)
            total_cultural_fit += r.evaluation_data.get("cultural_fit_score", 0)
            count += 1

    if count > 0:
        avg_technical = round(total_technical / count, 1)
        avg_communication = round(total_communication / count, 1)
        avg_problem_solving = round(total_problem_solving / count, 1)
        avg_cultural_fit = round(total_cultural_fit / count, 1)
        total_score = round(avg_technical + avg_communication + avg_problem_solving + avg_cultural_fit, 1)

        # 确定表现等级
        if total_score >= 32:
            level = "优秀"
        elif total_score >= 24:
            level = "良好"
        elif total_score >= 16:
            level = "合格"
        else:
            level = "待提升"
    else:
        avg_technical = avg_communication = avg_problem_solving = avg_cultural_fit = 0
        total_score = 0
        level = "未评估"

    return {
        "evaluation": {
            "technical_score": avg_technical,
            "communication_score": avg_communication,
            "problem_solving_score": avg_problem_solving,
            "cultural_fit_score": avg_cultural_fit,
            "total_score": total_score,
            "performance_level": level,
            "strengths": [],
            "weaknesses": [],
            "suggestions": []
        }
    }


@router.get("/sessions/{session_id}/download-pdf")
async def download_pdf_report(
    session_id: str,
    db: Session = Depends(get_digital_interviewer_db)
):
    """下载面试报告PDF"""
    from backend.models.db_models import InterviewRound, InterviewEvaluation
    from fastapi.responses import Response
    from datetime import datetime

    # 获取会话
    session = db.query(InterviewSession).filter_by(session_id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    # 获取轮次数据
    rounds = db.query(InterviewRound).filter_by(session_id=session.id).order_by(InterviewRound.round_number).all()

    # 获取评估数据
    evaluation = db.query(InterviewEvaluation).filter_by(session_id=session.id).first()

    # 如果没有评估，从轮次汇总
    if not evaluation:
        total_technical = total_communication = total_problem_solving = total_cultural_fit = 0
        count = 0
        for r in rounds:
            if r.evaluation_data:
                total_technical += r.evaluation_data.get("technical_score", 0)
                total_communication += r.evaluation_data.get("communication_score", 0)
                total_problem_solving += r.evaluation_data.get("problem_solving_score", 0)
                total_cultural_fit += r.evaluation_data.get("cultural_fit_score", 0)
                count += 1

        if count > 0:
            eval_data = {
                "technical_score": round(total_technical / count, 1),
                "communication_score": round(total_communication / count, 1),
                "problem_solving_score": round(total_problem_solving / count, 1),
                "cultural_fit_score": round(total_cultural_fit / count, 1),
            }
            eval_data["total_score"] = round(sum(eval_data.values()), 1)
        else:
            eval_data = {
                "technical_score": 0,
                "communication_score": 0,
                "problem_solving_score": 0,
                "cultural_fit_score": 0,
                "total_score": 0,
            }
    else:
        eval_data = {
            "technical_score": evaluation.technical_score or 0,
            "communication_score": evaluation.communication_score or 0,
            "problem_solving_score": evaluation.problem_solving_score or 0,
            "cultural_fit_score": evaluation.cultural_fit_score or 0,
            "total_score": evaluation.total_score or 0,
        }

    # 生成HTML内容
    rounds_html = ""
    for i, r in enumerate(rounds):
        eval_info = ""
        if r.evaluation_data:
            eval_info = f"""
            <p><strong>评分:</strong> 技术 {r.evaluation_data.get('technical_score', 0)}/10 |
            沟通 {r.evaluation_data.get('communication_score', 0)}/10 |
            问题解决 {r.evaluation_data.get('problem_solving_score', 0)}/10</p>
            """
        rounds_html += f"""
        <div class="round">
            <h3>第 {i + 1} 轮</h3>
            <p><strong>问题:</strong> {r.question}</p>
            <p><strong>回答:</strong> {r.answer or '未回答'}</p>
            {eval_info}
        </div>
        """

    interview_type_map = {
        "technical": "技术面试",
        "hr": "HR面试",
        "behavioral": "行为面试"
    }

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>面试报告</title>
        <style>
            @page {{ size: A4; margin: 2cm; }}
            body {{ font-family: 'SimSun', 'Microsoft YaHei', sans-serif; line-height: 1.6; color: #333; }}
            h1 {{ color: #1976d2; border-bottom: 2px solid #1976d2; padding-bottom: 10px; }}
            h2 {{ color: #333; margin-top: 30px; }}
            h3 {{ color: #1976d2; }}
            .info-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin: 20px 0; }}
            .info-item {{ padding: 10px; background: #f5f5f5; border-radius: 5px; }}
            .scores {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin: 20px 0; }}
            .score-item {{ text-align: center; padding: 15px; background: #e3f2fd; border-radius: 8px; }}
            .score-value {{ font-size: 24px; font-weight: bold; color: #1976d2; }}
            .total-score {{ text-align: center; padding: 20px; background: #1976d2; color: white; border-radius: 10px; margin: 20px 0; }}
            .total-score .value {{ font-size: 48px; font-weight: bold; }}
            .round {{ margin: 20px 0; padding: 15px; background: #fafafa; border-left: 4px solid #1976d2; border-radius: 0 8px 8px 0; }}
            .footer {{ margin-top: 40px; text-align: center; color: #666; font-size: 12px; }}
        </style>
    </head>
    <body>
        <h1>面试报告</h1>

        <h2>基本信息</h2>
        <div class="info-grid">
            <div class="info-item"><strong>候选人:</strong> {session.candidate_name or '未提供'}</div>
            <div class="info-item"><strong>面试官:</strong> {session.interviewer_name}</div>
            <div class="info-item"><strong>面试类型:</strong> {interview_type_map.get(session.interview_type, session.interview_type)}</div>
            <div class="info-item"><strong>面试时间:</strong> {session.created_at.strftime('%Y-%m-%d %H:%M')}</div>
        </div>

        <h2>综合评分</h2>
        <div class="total-score">
            <div class="value">{eval_data['total_score']}/40</div>
            <div>综合得分</div>
        </div>
        <div class="scores">
            <div class="score-item">
                <div class="score-value">{eval_data['technical_score']}</div>
                <div>技术能力</div>
            </div>
            <div class="score-item">
                <div class="score-value">{eval_data['communication_score']}</div>
                <div>沟通能力</div>
            </div>
            <div class="score-item">
                <div class="score-value">{eval_data['problem_solving_score']}</div>
                <div>问题解决</div>
            </div>
            <div class="score-item">
                <div class="score-value">{eval_data['cultural_fit_score']}</div>
                <div>文化匹配</div>
            </div>
        </div>

        <h2>面试详情</h2>
        {rounds_html if rounds_html else '<p>暂无面试轮次数据</p>'}

        <div class="footer">
            <p>报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>VividCrowd AI 数字面试官系统</p>
        </div>
    </body>
    </html>
    """

    # 尝试使用 WeasyPrint 生成 PDF
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=interview_report_{session_id}.pdf"
            }
        )
    except ImportError:
        # 如果没有安装 WeasyPrint，返回 HTML
        return Response(
            content=html_content.encode('utf-8'),
            media_type="text/html",
            headers={
                "Content-Disposition": f"attachment; filename=interview_report_{session_id}.html"
            }
        )


@router.websocket("/training/ws/{session_id}")
async def interview_websocket(websocket: WebSocket, session_id: str):
    """面试WebSocket端点 - 完整实现"""
    await websocket.accept()
    db = None
    orchestrator = None
    interview_active = False

    try:
        # 1. 获取数据库会话
        db = next(get_digital_interviewer_db())

        # 2. 初始化编排器（传递API Key）
        from backend.apps.digital_interviewer.services.interviewer_orchestrator import InterviewOrchestrator
        from backend.core.config import settings
        orchestrator = InterviewOrchestrator(
            db=db,
            api_key=settings.DASHSCOPE_API_KEY,
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"
        )

        # 3. 等待客户端发送开始信号
        await websocket.send_json({
            "type": "ready",
            "message": "连接成功，等待开始面试"
        })

        # 4. 面试循环
        while True:
            # 接收消息
            data = await websocket.receive_json()

            if data.get("type") == "start_interview":
                # 开始面试并推送第一个问题
                if not interview_active:
                    try:
                        # 获取面经配置（从客户端传递或从会话中获取）
                        experience_set_ids = data.get("experience_set_ids", [])
                        experience_mode = data.get("experience_mode", "none")

                        first_question = await orchestrator.start_interview(
                            session_id,
                            experience_set_ids=experience_set_ids,
                            experience_mode=experience_mode
                        )
                        interview_active = True
                        await websocket.send_json({
                            "type": "question",
                            "content": first_question["question"],
                            "round_number": first_question["round_number"],
                            "video_state": "speaking"
                        })
                    except Exception as e:
                        await websocket.send_json({
                            "type": "error",
                            "message": f"开始面试失败: {str(e)}"
                        })

            elif data.get("type") == "end_interview":
                # 结束面试
                if interview_active:
                    try:
                        final_eval = await orchestrator.end_interview(session_id)
                        await websocket.send_json({
                            "type": "interview_end",
                            "evaluation": final_eval,
                            "message": "面试已结束"
                        })
                        interview_active = False
                        break
                    except Exception as e:
                        await websocket.send_json({
                            "type": "error",
                            "message": f"结束面试失败: {str(e)}"
                        })

            elif data.get("type") == "answer" and interview_active:
                # 推送listening状态
                await websocket.send_json({
                    "type": "video_state",
                    "state": "listening"
                })

                # 处理回答
                try:
                    result = await orchestrator.process_answer(
                        session_id=session_id,
                        answer=data.get("content", "")
                    )

                    # 推送thinking状态
                    await websocket.send_json({
                        "type": "video_state",
                        "state": "thinking"
                    })

                    # 推送评估结果
                    await websocket.send_json({
                        "type": "evaluation",
                        "data": result.get("evaluation", {})
                    })

                    # 检查是否继续
                    if result.get("action") == "continue":
                        # 推送下一个问题
                        await websocket.send_json({
                            "type": "question",
                            "content": result["question"],
                            "round_number": result["round_number"],
                            "video_state": "speaking"
                        })
                    else:
                        # 面试结束
                        await websocket.send_json({
                            "type": "interview_end",
                            "evaluation": result.get("final_evaluation", {}),
                            "message": "面试已结束"
                        })
                        break

                except Exception as e:
                    await websocket.send_json({
                        "type": "error",
                        "message": f"处理回答失败: {str(e)}"
                    })

    except WebSocketDisconnect:
        # WebSocket断开时，更新会话状态为completed
        if db:
            try:
                session = db.query(InterviewSession).filter_by(session_id=session_id).first()
                if session and session.status == "in_progress":
                    from datetime import datetime
                    session.status = "completed"
                    session.completed_at = datetime.utcnow()
                    if session.started_at:
                        session.duration_seconds = int((session.completed_at - session.started_at).total_seconds())
                    db.commit()
            except Exception as e:
                print(f"更新会话状态失败: {e}")
    except Exception as e:
        try:
            await websocket.send_json({
                "type": "error",
                "message": f"WebSocket错误: {str(e)}"
            })
        except:
            pass
    finally:
        # 确保会话状态被更新
        if db:
            try:
                session = db.query(InterviewSession).filter_by(session_id=session_id).first()
                if session and session.status == "in_progress":
                    from datetime import datetime
                    session.status = "completed"
                    session.completed_at = datetime.utcnow()
                    if session.started_at:
                        session.duration_seconds = int((session.completed_at - session.started_at).total_seconds())
                    db.commit()
            except:
                pass
            db.close()


# ==================== 语音服务端点 ====================

@router.post("/audio/transcribe")
async def transcribe_audio_endpoint(file: UploadFile = File(...)):
    """音频转文字 (ASR) - 用于识别候选人的语音回答"""
    from backend.apps.digital_customer.services.audio_service import transcribe_audio
    from loguru import logger

    try:
        content = await file.read()
        extension = file.filename.split(".")[-1] if "." in file.filename else "wav"
        text = await transcribe_audio(content, extension)
        return {"text": text}
    except Exception as e:
        logger.error(f"语音识别失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/audio/synthesize")
async def synthesize_audio_endpoint(
    text: str = Form(...),
    voice: str = Form("longxiaochun")
):
    """文字转音频 (TTS) - 用于面试官语音提问"""
    from backend.apps.digital_customer.services.audio_service import synthesize_audio
    from fastapi.responses import Response
    from loguru import logger

    try:
        audio_data = await synthesize_audio(text, voice)
        return Response(content=audio_data, media_type="audio/mpeg")
    except Exception as e:
        logger.error(f"语音合成失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sessions")
async def get_interview_sessions(db: Session = Depends(get_digital_interviewer_db)):
    """获取所有面试会话列表"""
    sessions = db.query(InterviewSession).order_by(InterviewSession.created_at.desc()).all()
    return {"sessions": [session.to_dict() for session in sessions]}


# ==================== 面经管理端点 ====================

@router.post("/experience-sets/upload-stream")
async def upload_experience_set_stream(
    name: str = Form(...),
    interview_type: str = Form("technical"),
    description: str = Form(None),
    company: str = Form(None),
    position: str = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_digital_interviewer_db)
):
    """上传PDF面经并创建面经集（SSE流式返回进度）"""
    from fastapi.responses import StreamingResponse
    from backend.core.config import settings
    from backend.apps.digital_interviewer.services.experience_parser import ExperienceParser
    from loguru import logger
    import json as json_module

    # 先读取文件内容（在生成器外部读取，避免文件关闭问题）
    file_content = await file.read()
    filename = file.filename

    async def generate_progress():
        try:
            logger.info(f"开始上传面经: {filename}")

            # 验证文件类型
            if not filename.lower().endswith('.pdf'):
                yield f"data: {json_module.dumps({'type': 'error', 'message': '只支持PDF文件'})}\n\n"
                return

            # 保存文件
            file_path = os.path.join(EXPERIENCE_DIR, filename)
            with open(file_path, 'wb') as f:
                f.write(file_content)

            yield f"data: {json_module.dumps({'type': 'progress', 'current': 0, 'total': 100, 'message': '文件已保存，开始解析...'})}\n\n"

            # 创建解析器
            parser = ExperienceParser(
                api_key=settings.DASHSCOPE_API_KEY,
                base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"
            )

            # 流式解析
            parse_result = None
            async for progress in parser.parse_pdf_stream(file_path):
                if progress["type"] == "complete":
                    parse_result = progress["result"]
                yield f"data: {json_module.dumps(progress)}\n\n"

            if not parse_result:
                yield f"data: {json_module.dumps({'type': 'error', 'message': '解析失败'})}\n\n"
                return

            # 保存到数据库
            yield f"data: {json_module.dumps({'type': 'progress', 'current': 100, 'total': 100, 'message': '正在保存到数据库...'})}\n\n"

            experience_set = InterviewExperienceSet(
                name=name,
                description=description or parse_result.get("title"),
                source_filename=filename,
                company=company or parse_result.get("company"),
                position=position or parse_result.get("position"),
                interview_type=interview_type or parse_result.get("interview_type", "technical"),
                question_count=len(parse_result.get("questions", [])),
                is_active=True
            )
            db.add(experience_set)
            db.flush()

            # 保存问题
            for q in parse_result.get("questions", []):
                knowledge = InterviewKnowledge(
                    interview_type=interview_type,
                    category=q.get("category"),
                    content=q.get("question", ""),
                    difficulty_level=q.get("difficulty"),
                    source_filename=filename,
                    experience_set_id=experience_set.id,
                    question_text=q.get("question", ""),
                    reference_answer=q.get("answer"),
                    tags=json_module.dumps(q.get("tags")) if q.get("tags") else None
                )
                db.add(knowledge)

            db.commit()
            db.refresh(experience_set)

            yield f"data: {json_module.dumps({'type': 'done', 'message': '上传成功', 'experience_set': experience_set.to_dict()})}\n\n"

        except Exception as e:
            logger.error(f"上传面经失败: {e}")
            import traceback
            traceback.print_exc()
            yield f"data: {json_module.dumps({'type': 'error', 'message': str(e)})}\n\n"

    return StreamingResponse(
        generate_progress(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )


@router.post("/experience-sets/upload")
async def upload_experience_set(
    name: str = Form(...),
    interview_type: str = Form("technical"),
    description: str = Form(None),
    company: str = Form(None),
    position: str = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_digital_interviewer_db)
):
    """上传PDF面经并创建面经集"""
    from loguru import logger
    from backend.core.config import settings

    try:
        logger.info(f"开始上传面经: {file.filename}")

        # 验证文件类型
        if not file.filename.lower().endswith('.pdf'):
            raise HTTPException(status_code=400, detail="只支持PDF文件")

        # 保存文件
        content = await file.read()
        file_path = os.path.join(EXPERIENCE_DIR, file.filename)
        with open(file_path, 'wb') as f:
            f.write(content)
        logger.info(f"文件已保存到: {file_path}")

        # 创建面经集
        service = ExperienceService(
            db=db,
            api_key=settings.DASHSCOPE_API_KEY,
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"
        )

        experience_set = await service.create_experience_set(
            name=name,
            file_path=file_path,
            interview_type=interview_type,
            description=description,
            company=company,
            position=position
        )

        return {
            "message": "上传成功",
            "experience_set": experience_set.to_dict()
        }

    except Exception as e:
        logger.error(f"上传面经失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/experience-sets")
async def list_experience_sets(
    interview_type: str = None,
    db: Session = Depends(get_digital_interviewer_db)
):
    """获取面经集列表"""
    service = ExperienceService(db=db)
    sets = service.get_experience_sets(interview_type=interview_type)
    return {"experience_sets": [s.to_dict() for s in sets]}


@router.get("/experience-sets/{set_id}")
async def get_experience_set(
    set_id: int,
    db: Session = Depends(get_digital_interviewer_db)
):
    """获取单个面经集详情"""
    service = ExperienceService(db=db)
    experience_set = service.get_experience_set(set_id)
    if not experience_set:
        raise HTTPException(status_code=404, detail="面经集不存在")
    return {"experience_set": experience_set.to_dict()}


@router.get("/experience-sets/{set_id}/questions")
async def get_experience_questions(
    set_id: int,
    db: Session = Depends(get_digital_interviewer_db)
):
    """获取面经集中的问题"""
    service = ExperienceService(db=db)
    questions = service.get_questions_by_set(set_id)
    return {"questions": [q.to_dict() for q in questions]}


@router.delete("/experience-sets/{set_id}")
async def delete_experience_set(
    set_id: int,
    db: Session = Depends(get_digital_interviewer_db)
):
    """删除面经集"""
    service = ExperienceService(db=db)
    success = service.delete_experience_set(set_id)
    if not success:
        raise HTTPException(status_code=404, detail="面经集不存在")
    return {"message": "删除成功"}


# ==================== 简历管理API ====================

@router.post("/resumes/upload")
async def upload_resume(
    file: UploadFile = File(...),
    candidate_name: str = Form(None),
    candidate_id: str = Form(None),
    db: Session = Depends(get_digital_interviewer_db)
):
    """上传并解析简历"""
    from loguru import logger
    from backend.models.db_models import CandidateResume, ResumeAnalysis
    from backend.apps.digital_interviewer.services.resume_parser import ResumeParser

    try:
        logger.info(f"📤 开始上传简历: {file.filename}")

        # 检查文件格式
        file_ext = file.filename.split('.')[-1].lower()
        if file_ext not in ['pdf', 'docx', 'doc', 'jpg', 'jpeg', 'png']:
            raise HTTPException(status_code=400, detail="不支持的文件格式")

        # 保存文件
        file_path = os.path.join(RESUME_DIR, f"{uuid.uuid4()}_{file.filename}")
        content = await file.read()
        with open(file_path, 'wb') as f:
            f.write(content)

        logger.info(f"💾 文件已保存: {file_path}")

        # 解析简历
        parser = ResumeParser()
        parse_result = parser.parse_resume(file_path)

        if not parse_result['success']:
            raise HTTPException(status_code=400, detail=parse_result['error'])

        # 创建简历记录
        resume = CandidateResume(
            candidate_id=candidate_id,
            candidate_name=candidate_name,
            file_path=file_path,
            file_type=file_ext,
            file_hash=parse_result['file_hash'],
            parse_status='completed',
            parsed_data=parse_result['structured_data']
        )

        # 评估简历质量
        quality_result = parser.evaluate_quality(
            parse_result['structured_data'],
            parse_result['raw_text']
        )
        resume.quality_score = quality_result['quality_score']
        resume.completeness_score = quality_result['completeness_score']
        resume.professionalism_score = quality_result['professionalism_score']

        db.add(resume)
        db.flush()

        # 创建分析记录
        tags = parser.extract_tags(parse_result['structured_data'])
        work_years = parser.calculate_work_years(parse_result['structured_data'])

        analysis = ResumeAnalysis(
            resume_id=resume.id,
            contact_info=parse_result['structured_data'].get('contact'),
            education=parse_result['structured_data'].get('education'),
            work_experience=parse_result['structured_data'].get('work_experience'),
            project_experience=parse_result['structured_data'].get('projects'),
            skills=parse_result['structured_data'].get('skills'),
            certifications=parse_result['structured_data'].get('certifications'),
            total_work_years=work_years,
            skill_tags=tags['skill_tags'],
            industry_tags=tags['industry_tags'],
            position_tags=tags['position_tags'],
            quality_issues=quality_result['issues'],
            improvement_suggestions=quality_result['suggestions'],
            risk_flags=quality_result['risk_flags']
        )

        db.add(analysis)
        db.commit()
        db.refresh(resume)

        logger.info(f"✅ 简历上传成功: ID={resume.id}")

        return {
            "message": "简历上传成功",
            "resume_id": resume.id,
            "resume": resume.to_dict(),
            "quality_score": resume.quality_score
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"上传简历失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/resumes")
async def list_resumes(
    skip: int = 0,
    limit: int = 20,
    db: Session = Depends(get_digital_interviewer_db)
):
    """获取简历列表"""
    from backend.models.db_models import CandidateResume

    resumes = db.query(CandidateResume).offset(skip).limit(limit).all()
    total = db.query(CandidateResume).count()

    return {
        "resumes": [resume.to_dict() for resume in resumes],
        "total": total,
        "skip": skip,
        "limit": limit
    }


@router.get("/resumes/{resume_id}")
async def get_resume(
    resume_id: int,
    db: Session = Depends(get_digital_interviewer_db)
):
    """获取简历详情"""
    from backend.models.db_models import CandidateResume, ResumeAnalysis

    resume = db.query(CandidateResume).filter(CandidateResume.id == resume_id).first()
    if not resume:
        raise HTTPException(status_code=404, detail="简历不存在")

    analysis = db.query(ResumeAnalysis).filter(ResumeAnalysis.resume_id == resume_id).first()

    return {
        "resume": resume.to_dict(),
        "analysis": analysis.to_dict() if analysis else None
    }


@router.delete("/resumes/{resume_id}")
async def delete_resume(
    resume_id: int,
    db: Session = Depends(get_digital_interviewer_db)
):
    """删除简历"""
    from backend.models.db_models import CandidateResume

    resume = db.query(CandidateResume).filter(CandidateResume.id == resume_id).first()
    if not resume:
        raise HTTPException(status_code=404, detail="简历不存在")

    # 删除文件
    if os.path.exists(resume.file_path):
        os.remove(resume.file_path)

    db.delete(resume)
    db.commit()

    return {"message": "删除成功"}


# ==================== 职位管理API ====================

@router.post("/jobs")
async def create_job(
    title: str = Form(...),
    department: str = Form(None),
    requirements: str = Form(None),
    skills_required: str = Form(None),
    education_required: str = Form(None),
    experience_years_min: int = Form(None),
    job_type: str = Form(None),
    db: Session = Depends(get_digital_interviewer_db)
):
    """创建职位"""
    from backend.models.db_models import JobPosition
    import json

    try:
        skills_list = json.loads(skills_required) if skills_required else []
    except:
        skills_list = []

    job = JobPosition(
        title=title,
        department=department,
        requirements=requirements,
        skills_required=skills_list,
        education_required=education_required,
        experience_years_min=experience_years_min,
        job_type=job_type,
        is_active=True
    )

    db.add(job)
    db.commit()
    db.refresh(job)

    return {"message": "职位创建成功", "job": job.to_dict()}


@router.get("/jobs")
async def list_jobs(
    is_active: bool = None,
    db: Session = Depends(get_digital_interviewer_db)
):
    """获取职位列表"""
    from backend.models.db_models import JobPosition

    query = db.query(JobPosition)
    if is_active is not None:
        query = query.filter(JobPosition.is_active == is_active)

    jobs = query.all()
    return {"jobs": [job.to_dict() for job in jobs]}


# ==================== 简历匹配API ====================

@router.post("/resumes/{resume_id}/match")
async def match_resume_to_jobs(
    resume_id: int,
    job_ids: List[int] = None,
    db: Session = Depends(get_digital_interviewer_db)
):
    """将简历与职位进行匹配"""
    from backend.apps.digital_interviewer.services.resume_matcher import ResumeMatcher

    try:
        matcher = ResumeMatcher(db)
        results = matcher.batch_match(resume_id, job_ids)

        return {
            "message": "匹配完成",
            "resume_id": resume_id,
            "matches": results
        }

    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== GDPR合规API ====================

@router.get("/candidates/{candidate_id}/export")
async def export_candidate_data(
    candidate_id: str,
    db: Session = Depends(get_digital_interviewer_db)
):
    """导出候选人的所有数据（GDPR合规）"""
    from backend.models.db_models import (
        CandidateResume, ResumeAnalysis, InterviewSession,
        InterviewRound, InterviewEvaluation
    )
    import json

    try:
        # 收集所有相关数据
        export_data = {
            "candidate_id": candidate_id,
            "export_time": datetime.utcnow().isoformat(),
            "resumes": [],
            "interviews": []
        }

        # 简历数据
        resumes = db.query(CandidateResume).filter(
            CandidateResume.candidate_id == candidate_id
        ).all()

        for resume in resumes:
            resume_data = resume.to_dict()
            analysis = db.query(ResumeAnalysis).filter(
                ResumeAnalysis.resume_id == resume.id
            ).first()
            if analysis:
                resume_data['analysis'] = analysis.to_dict()
            export_data['resumes'].append(resume_data)

        return JSONResponse(content=export_data)

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/candidates/{candidate_id}/data")
async def delete_candidate_data(
    candidate_id: str,
    db: Session = Depends(get_digital_interviewer_db)
):
    """删除候选人的所有数据（GDPR合规）"""
    from backend.models.db_models import CandidateResume, InterviewSession

    try:
        # 删除简历（级联删除分析数据）
        resumes = db.query(CandidateResume).filter(
            CandidateResume.candidate_id == candidate_id
        ).all()

        deleted_count = 0
        for resume in resumes:
            # 删除文件
            if os.path.exists(resume.file_path):
                os.remove(resume.file_path)
            db.delete(resume)
            deleted_count += 1

        # 删除面试会话（级联删除轮次和评估数据）
        sessions = db.query(InterviewSession).filter(
            InterviewSession.candidate_id == candidate_id
        ).all()

        for session in sessions:
            db.delete(session)
            deleted_count += 1

        db.commit()

        return {
            "message": "候选人数据已删除",
            "candidate_id": candidate_id,
            "deleted_items": deleted_count
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 评分模板管理API ====================

@router.post("/scoring-templates")
async def create_scoring_template(
    name: str = Form(...),
    description: str = Form(None),
    job_type: str = Form(None),
    technical_weight: int = Form(25),
    communication_weight: int = Form(15),
    problem_solving_weight: int = Form(20),
    cultural_fit_weight: int = Form(10),
    innovation_weight: int = Form(10),
    teamwork_weight: int = Form(10),
    stress_handling_weight: int = Form(5),
    learning_ability_weight: int = Form(5),
    db: Session = Depends(get_digital_interviewer_db)
):
    """创建评分模板"""
    from backend.models.db_models import ScoringTemplate

    # 验证权重总和
    total_weight = (technical_weight + communication_weight + problem_solving_weight +
                   cultural_fit_weight + innovation_weight + teamwork_weight +
                   stress_handling_weight + learning_ability_weight)

    if total_weight != 100:
        raise HTTPException(status_code=400, detail=f"权重总和必须为100，当前为{total_weight}")

    template = ScoringTemplate(
        name=name,
        description=description,
        job_type=job_type,
        technical_weight=technical_weight,
        communication_weight=communication_weight,
        problem_solving_weight=problem_solving_weight,
        cultural_fit_weight=cultural_fit_weight,
        innovation_weight=innovation_weight,
        teamwork_weight=teamwork_weight,
        stress_handling_weight=stress_handling_weight,
        learning_ability_weight=learning_ability_weight
    )

    db.add(template)
    db.commit()
    db.refresh(template)

    return {"message": "评分模板创建成功", "template": template.to_dict()}


@router.get("/scoring-templates")
async def list_scoring_templates(
    job_type: str = None,
    db: Session = Depends(get_digital_interviewer_db)
):
    """获取评分模板列表"""
    from backend.models.db_models import ScoringTemplate

    query = db.query(ScoringTemplate).filter(ScoringTemplate.is_active == True)
    if job_type:
        query = query.filter(ScoringTemplate.job_type == job_type)

    templates = query.all()
    return {"templates": [t.to_dict() for t in templates]}


# ==================== 面试模板管理API ====================

@router.post("/interview-templates")
async def create_interview_template(
    name: str = Form(...),
    description: str = Form(None),
    job_type: str = Form(None),
    max_rounds: int = Form(5),
    difficulty_level: str = Form("medium"),
    experience_set_ids: str = Form(None),
    scoring_template_id: int = Form(None),
    db: Session = Depends(get_digital_interviewer_db)
):
    """创建面试模板"""
    from backend.models.db_models import InterviewTemplate
    import json

    try:
        exp_ids = json.loads(experience_set_ids) if experience_set_ids else []
    except:
        exp_ids = []

    template = InterviewTemplate(
        name=name,
        description=description,
        job_type=job_type,
        max_rounds=max_rounds,
        difficulty_level=difficulty_level,
        experience_set_ids=exp_ids,
        scoring_template_id=scoring_template_id
    )

    db.add(template)
    db.commit()
    db.refresh(template)

    return {"message": "面试模板创建成功", "template": template.to_dict()}


# ==================== 候选人管理API ====================

@router.post("/candidates/batch-import")
async def batch_import_candidates(
    file: UploadFile = File(...),
    db: Session = Depends(get_digital_interviewer_db)
):
    """批量导入候选人（Excel/CSV）"""
    from backend.models.db_models import Candidate
    import pandas as pd
    import uuid
    import io

    try:
        # 检查文件格式
        file_ext = file.filename.split('.')[-1].lower()
        if file_ext not in ['xlsx', 'xls', 'csv']:
            raise HTTPException(status_code=400, detail="仅支持Excel或CSV格式")

        # 读取文件
        content = await file.read()
        if file_ext == 'csv':
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))

        # 验证必需列
        required_columns = ['姓名']
        for col in required_columns:
            if col not in df.columns:
                raise HTTPException(status_code=400, detail=f"缺少必需列: {col}")

        success_count = 0
        failed_count = 0
        errors = []

        for index, row in df.iterrows():
            try:
                candidate = Candidate(
                    candidate_id=str(uuid.uuid4()),
                    name=str(row['姓名']),
                    email=str(row.get('邮箱', '')),
                    phone=str(row.get('电话', '')),
                    position=str(row.get('职位', '')),
                    source='batch_import'
                )
                db.add(candidate)
                success_count += 1
            except Exception as e:
                failed_count += 1
                errors.append(f"第{index+2}行: {str(e)}")

        db.commit()

        return {
            "message": "批量导入完成",
            "success_count": success_count,
            "failed_count": failed_count,
            "errors": errors[:10]
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/interviews/export")
async def export_interviews(
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_digital_interviewer_db)
):
    """导出面试结果为Excel"""
    from backend.models.db_models import InterviewSession, InterviewEvaluation
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    import io
    from datetime import datetime

    try:
        # 查询面试会话
        query = db.query(InterviewSession).filter(InterviewSession.status == "completed")

        if start_date:
            query = query.filter(InterviewSession.created_at >= start_date)
        if end_date:
            query = query.filter(InterviewSession.created_at <= end_date)

        sessions = query.all()

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "面试结果"

        # 设置表头
        headers = [
            "会话ID", "候选人姓名", "面试类型", "面试官",
            "开始时间", "完成时间", "总轮数", "状态",
            "技术能力", "沟通表达", "问题解决", "文化匹配",
            "创新能力", "团队协作", "压力应对", "学习能力", "总分"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 填充数据
        for row_idx, session in enumerate(sessions, 2):
            evaluation = db.query(InterviewEvaluation).filter(
                InterviewEvaluation.session_id == session.id
            ).first()

            ws.cell(row=row_idx, column=1, value=session.session_id)
            ws.cell(row=row_idx, column=2, value=session.candidate_name or "")
            ws.cell(row=row_idx, column=3, value=session.interview_type)
            ws.cell(row=row_idx, column=4, value=session.interviewer_name or "")
            ws.cell(row=row_idx, column=5, value=session.started_at.strftime("%Y-%m-%d %H:%M") if session.started_at else "")
            ws.cell(row=row_idx, column=6, value=session.completed_at.strftime("%Y-%m-%d %H:%M") if session.completed_at else "")
            ws.cell(row=row_idx, column=7, value=session.total_rounds)
            ws.cell(row=row_idx, column=8, value=session.status)

            if evaluation:
                ws.cell(row=row_idx, column=9, value=evaluation.technical_score or 0)
                ws.cell(row=row_idx, column=10, value=evaluation.communication_score or 0)
                ws.cell(row=row_idx, column=11, value=evaluation.problem_solving_score or 0)
                ws.cell(row=row_idx, column=12, value=evaluation.cultural_fit_score or 0)
                ws.cell(row=row_idx, column=13, value=evaluation.innovation_score or 0)
                ws.cell(row=row_idx, column=14, value=evaluation.teamwork_score or 0)
                ws.cell(row=row_idx, column=15, value=evaluation.stress_handling_score or 0)
                ws.cell(row=row_idx, column=16, value=evaluation.learning_ability_score or 0)
                ws.cell(row=row_idx, column=17, value=evaluation.total_score or 0)

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 返回文件
        from fastapi.responses import StreamingResponse
        filename = f"interview_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
